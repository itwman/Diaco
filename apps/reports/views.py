"""
Diaco MES - Reports Views
============================
گزارش‌گیری: تولید، مصرف مواد، توقفات، تکمیل نخ v2.0، زنجیره.
خروجی Excel با openpyxl.
"""
import json
import jdatetime
from datetime import date, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, Avg, Q, F, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

from apps.blowroom.models import Batch as BlowroomBatch
from apps.carding.models import Production as CardingProd
from apps.passage.models import Production as PassageProd
from apps.finisher.models import Production as FinisherProd
from apps.spinning.models import Production as SpinningProd
from apps.dyeing.models import Batch as DyeingBatch, ChemicalUsage
from apps.inventory.models import FiberStock, DyeStock, ChemicalStock
from apps.core.models import ProductionLine
from apps.maintenance.models import WorkOrder, DowntimeLog, Schedule
from apps.winding.models import Production as WindingProd
from apps.tfo.models import Production as TFOProd
from apps.heatset.models import Batch as HeatsetBatch


# ═══════════════════════════════════════════════════════════════
# ابزارهای مشترک
# ═══════════════════════════════════════════════════════════════

def _parse_date_range(request):
    """استخراج بازه تاریخ از GET params."""
    today = date.today()
    date_from_str = request.GET.get('from', '')
    date_to_str   = request.GET.get('to', '')
    try:
        date_from = date.fromisoformat(date_from_str)
    except (ValueError, TypeError):
        date_from = today - timedelta(days=7)
    try:
        date_to = date.fromisoformat(date_to_str)
    except (ValueError, TypeError):
        date_to = today
    return date_from, date_to


def _parse_line(request):
    """استخراج خط تولید از GET params."""
    line_id = request.GET.get('line')
    if line_id:
        try:
            return ProductionLine.objects.get(pk=line_id)
        except ProductionLine.DoesNotExist:
            pass
    return None


def _filter_by_line(qs, line, field='production_line'):
    if line:
        return qs.filter(**{field: line})
    return qs


def _report_base_context(request):
    return {
        'lines':         ProductionLine.objects.filter(status='active'),
        'selected_line': _parse_line(request),
    }


def _jstr(d):
    """تبدیل date به رشته شمسی."""
    return jdatetime.date.fromgregorian(date=d).strftime('%Y/%m/%d')


# ═══════════════════════════════════════════════════════════════
# 9.1 گزارش تولید روزانه
# ═══════════════════════════════════════════════════════════════

@login_required
def production_daily(request):
    """گزارش تولید روزانه تمام بخش‌ها."""
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']

    sections = []
    for label, Model, weight_field in [
        ('حلاجی',   BlowroomBatch, 'output_weight'),
        ('کاردینگ', CardingProd,   'output_weight'),
        ('پاساژ',   PassageProd,   'output_weight_gperm'),
        ('فینیشر',  FinisherProd,  'output_weight'),
        ('رینگ',    SpinningProd,  'output_weight'),
    ]:
        qs  = _filter_by_line(
            Model.objects.filter(production_date__range=(date_from, date_to)), line
        )
        agg = qs.aggregate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='completed')),
            total_weight=Sum(weight_field),
        )
        sections.append({
            'label': label,
            'total': agg['total'],
            'completed': agg['completed'],
            'total_weight': agg['total_weight'] or 0,
        })

    # بوبین‌پیچی
    wd_qs  = _filter_by_line(WindingProd.objects.filter(production_date__range=(date_from, date_to)), line)
    wd_agg = wd_qs.aggregate(total=Count('id'), completed=Count('id', filter=Q(status='completed')), total_weight=Sum('output_weight_kg'))
    sections.append({'label': 'بوبین‌پیچی', 'total': wd_agg['total'], 'completed': wd_agg['completed'], 'total_weight': wd_agg['total_weight'] or 0})

    # دولاتابی
    tfo_qs  = _filter_by_line(TFOProd.objects.filter(production_date__range=(date_from, date_to)), line)
    tfo_agg = tfo_qs.aggregate(total=Count('id'), completed=Count('id', filter=Q(status='completed')), total_weight=Sum('output_weight_kg'))
    sections.append({'label': 'دولاتابی', 'total': tfo_agg['total'], 'completed': tfo_agg['completed'], 'total_weight': tfo_agg['total_weight'] or 0})

    # هیت‌ست
    hs_qs  = _filter_by_line(HeatsetBatch.objects.filter(production_date__range=(date_from, date_to)), line)
    hs_agg = hs_qs.aggregate(total=Count('id'), completed=Count('id', filter=Q(quality_result='pass')), total_weight=Sum('batch_weight_kg'))
    sections.append({'label': 'هیت‌ست', 'total': hs_agg['total'], 'completed': hs_agg['completed'], 'total_weight': hs_agg['total_weight'] or 0})

    # رنگرزی
    dye_qs  = DyeingBatch.objects.filter(production_date__range=(date_from, date_to))
    if line:
        dye_qs = dye_qs.filter(machine__production_line=line)
    dye_agg = dye_qs.aggregate(total=Count('id'), completed=Count('id', filter=Q(status='completed')), total_weight=Sum('fiber_weight'))
    sections.append({'label': 'رنگرزی', 'total': dye_agg['total'], 'completed': dye_agg['completed'], 'total_weight': dye_agg['total_weight'] or 0})

    # نمودار روزانه رینگ
    chart_data = []
    current = date_from
    while current <= date_to:
        sp_qs = _filter_by_line(SpinningProd.objects.filter(production_date=current, status='completed'), line)
        weight = sp_qs.aggregate(w=Sum('output_weight'))['w'] or 0
        chart_data.append({'date': _jstr(current), 'weight': float(weight)})
        current += timedelta(days=1)

    ctx.update({
        'sections': sections,
        'date_from': date_from,
        'date_to': date_to,
        'chart_data': json.dumps(chart_data),
        'page_title': 'گزارش تولید روزانه',
    })
    return render(request, 'reports/production_daily.html', ctx)


# ═══════════════════════════════════════════════════════════════
# 9.2 گزارش مصرف مواد
# ═══════════════════════════════════════════════════════════════

@login_required
def material_consumption(request):
    date_from, date_to = _parse_date_range(request)
    from apps.blowroom.models import BatchInput
    fiber_usage = BatchInput.objects.filter(
        batch__production_date__range=(date_from, date_to)
    ).values('fiber_stock__category__name').annotate(total_used=Sum('weight_used')).order_by('-total_used')

    chemical_usage = ChemicalUsage.objects.filter(
        dyeing_batch__production_date__range=(date_from, date_to)
    ).values('material_type').annotate(total_used=Sum('quantity_used'), count=Count('id')).order_by('-total_used')

    fiber_stock_total = FiberStock.objects.filter(status='available').aggregate(total=Sum('current_weight'))['total'] or 0

    chart_data = []
    current = date_from
    while current <= date_to:
        w = BatchInput.objects.filter(batch__production_date=current).aggregate(s=Sum('weight_used'))['s'] or 0
        chart_data.append({'date': _jstr(current), 'weight': float(w)})
        current += timedelta(days=1)

    return render(request, 'reports/material_consumption.html', {
        'fiber_usage': fiber_usage,
        'chemical_usage': chemical_usage,
        'fiber_stock_total': fiber_stock_total,
        'dye_stock_count': DyeStock.objects.filter(status='available').count(),
        'chem_stock_count': ChemicalStock.objects.filter(status='available').count(),
        'date_from': date_from,
        'date_to': date_to,
        'chart_data': json.dumps(chart_data),
        'page_title': 'گزارش مصرف مواد',
    })


# ═══════════════════════════════════════════════════════════════
# 9.3 گزارش توقفات
# ═══════════════════════════════════════════════════════════════

@login_required
def downtime_report(request):
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']
    dt_base = _filter_by_line(DowntimeLog.objects.filter(start_time__date__range=(date_from, date_to)), line)

    by_reason  = dt_base.values('reason_category').annotate(count=Count('id'), total_min=Sum('duration_min'), total_loss=Sum('production_loss')).order_by('-total_min')
    by_machine = dt_base.values('machine__code').annotate(count=Count('id'), total_min=Sum('duration_min')).order_by('-total_min')[:10]
    wo_stats   = WorkOrder.objects.filter(created_at__date__range=(date_from, date_to)).aggregate(
        total=Count('id'), completed=Count('id', filter=Q(status='completed')),
        total_downtime=Sum('downtime_min'), total_cost=Sum(F('cost_parts') + F('cost_labor')),
    )

    from django.utils import timezone
    pm_overdue = Schedule.objects.filter(is_active=True, next_due_at__lt=timezone.now()).select_related('machine').order_by('next_due_at')[:10]

    reason_labels = {'mechanical': 'مکانیکی', 'electrical': 'برقی', 'material': 'مواد',
                     'operator': 'اپراتور', 'quality': 'کیفیت', 'planned': 'برنامه‌ریزی', 'other': 'سایر'}
    pie_data = [{'label': reason_labels.get(r['reason_category'], r['reason_category']), 'value': r['total_min'] or 0} for r in by_reason]

    ctx.update({'by_reason': by_reason, 'by_machine': by_machine, 'wo_stats': wo_stats,
                'pm_overdue': pm_overdue, 'reason_labels': reason_labels,
                'date_from': date_from, 'date_to': date_to,
                'pie_data': json.dumps(pie_data), 'page_title': 'گزارش توقفات'})
    return render(request, 'reports/downtime_report.html', ctx)


# ═══════════════════════════════════════════════════════════════
# R1.1 گزارش ضایعات
# ═══════════════════════════════════════════════════════════════

@login_required
def waste_report(request):
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']
    sections = []
    for label, Model, input_f, output_f, waste_f in [
        ('حلاجی',  BlowroomBatch, 'total_input_weight', 'output_weight', 'waste_weight'),
        ('کاردینگ', CardingProd,  'input_weight',       'output_weight', 'waste_weight'),
        ('فینیشر',  FinisherProd, 'input_weight',       'output_weight', None),
        ('رینگ',    SpinningProd, 'input_weight',       'output_weight', None),
    ]:
        qs = _filter_by_line(Model.objects.filter(production_date__range=(date_from, date_to), status='completed'), line)
        agg_kwargs = {'total_input': Sum(input_f), 'total_output': Sum(output_f), 'batch_count': Count('id')}
        if waste_f:
            agg_kwargs['total_waste'] = Sum(waste_f)
        agg = qs.aggregate(**agg_kwargs)
        total_in  = float(agg['total_input']  or 0)
        total_out = float(agg['total_output'] or 0)
        total_waste = float(agg.get('total_waste') or (total_in - total_out if not waste_f else 0))
        sections.append({'label': label, 'batch_count': agg['batch_count'],
                         'total_input': total_in, 'total_output': total_out,
                         'total_waste': total_waste,
                         'waste_pct': round(total_waste / total_in * 100, 2) if total_in > 0 else 0})

    # بوبین‌پیچی ضایعات
    wd_qs = _filter_by_line(WindingProd.objects.filter(production_date__range=(date_from, date_to), status='completed'), line)
    wd_w  = wd_qs.aggregate(i=Sum('input_weight_kg'), o=Sum('output_weight_kg'), w=Sum('waste_weight_kg'), c=Count('id'))
    wi = float(wd_w['i'] or 0); wo = float(wd_w['o'] or 0); ww = float(wd_w['w'] or 0)
    sections.append({'label': 'بوبین‌پیچی', 'batch_count': wd_w['c'], 'total_input': wi,
                     'total_output': wo, 'total_waste': ww,
                     'waste_pct': round(ww / wi * 100, 2) if wi > 0 else 0})

    # دولاتابی ضایعات
    tfo_qs = _filter_by_line(TFOProd.objects.filter(production_date__range=(date_from, date_to), status='completed'), line)
    tfo_w  = tfo_qs.aggregate(i=Sum('input_weight_kg'), o=Sum('output_weight_kg'), w=Sum('waste_weight_kg'), c=Count('id'))
    ti = float(tfo_w['i'] or 0); to_ = float(tfo_w['o'] or 0); tw = float(tfo_w['w'] or 0)
    sections.append({'label': 'دولاتابی', 'batch_count': tfo_w['c'], 'total_input': ti,
                     'total_output': to_, 'total_waste': tw,
                     'waste_pct': round(tw / ti * 100, 2) if ti > 0 else 0})

    by_machine = []
    for label, Model in [('حلاجی', BlowroomBatch), ('کاردینگ', CardingProd)]:
        qs = _filter_by_line(Model.objects.filter(production_date__range=(date_from, date_to), status='completed'), line)
        inp_f = 'total_input_weight' if label == 'حلاجی' else 'input_weight'
        for m in qs.values('machine__code').annotate(tw=Sum('waste_weight'), ti=Sum(inp_f), c=Count('id')).order_by('-tw'):
            inp = float(m['ti'] or 0); wst = float(m['tw'] or 0)
            by_machine.append({'section': label, 'machine': m['machine__code'], 'waste_kg': wst,
                               'waste_pct': round(wst / inp * 100, 2) if inp > 0 else 0, 'batch_count': m['c']})
    by_machine.sort(key=lambda x: x['waste_kg'], reverse=True)

    trend_data = []
    current = date_from
    while current <= date_to:
        bl_qs = _filter_by_line(BlowroomBatch.objects.filter(production_date=current, status='completed'), line)
        agg = bl_qs.aggregate(w=Sum('waste_weight'), i=Sum('total_input_weight'))
        inp = float(agg['i'] or 0); wst = float(agg['w'] or 0)
        trend_data.append({'date': _jstr(current), 'waste_pct': round(wst / inp * 100, 2) if inp > 0 else 0, 'waste_kg': wst})
        current += timedelta(days=1)

    ctx.update({'sections': sections, 'by_machine': by_machine[:15], 'date_from': date_from, 'date_to': date_to,
                'chart_data': json.dumps([{'label': s['label'], 'value': s['waste_pct']} for s in sections]),
                'trend_data': json.dumps(trend_data), 'page_title': 'گزارش ضایعات'})
    return render(request, 'reports/waste_report.html', ctx)


# ═══════════════════════════════════════════════════════════════
# R1.2 گزارش OEE
# ═══════════════════════════════════════════════════════════════

@login_required
def oee_report(request):
    """
    F.3 — OEE گسترش‌یافته: رینگ + بوبین‌پیچی + دولاتابی + هیت‌ست.
    OEE = Availability × Performance × Quality
    برای WD/TFO: Quality از efficiency_pct تخمین زده می‌شود.
    برای HS: Quality = pass_rate.
    """
    from apps.core.models import Machine
    from apps.ai_ready.utils import calculate_oee
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']

    # ── رینگ (روش اصلی با calculate_oee) ───────────────────────
    ring_machines = Machine.objects.filter(status='active', machine_type='ring')
    if line:
        ring_machines = ring_machines.filter(production_line=line)

    machine_oee = []
    for m in ring_machines:
        oee_sum = {'availability': 0, 'performance': 0, 'quality': 0, 'oee': 0}
        valid_days = 0
        current = date_from
        while current <= date_to:
            day_oee = calculate_oee(m.id, current)
            if day_oee['batch_count'] > 0:
                for k in oee_sum:
                    oee_sum[k] += day_oee[k]
                valid_days += 1
            current += timedelta(days=1)
        if valid_days > 0:
            for k in oee_sum:
                oee_sum[k] = round(oee_sum[k] / valid_days, 1)
        sp_agg = SpinningProd.objects.filter(
            machine=m, production_date__range=(date_from, date_to), status='completed'
        ).aggregate(
            avg_eff=Avg('efficiency_pct'), total_breakage=Sum('breakage_count'),
            total_output=Sum('output_weight'), batch_count=Count('id')
        )
        machine_oee.append({
            'code': m.code, 'name': m.name, 'section': 'رینگ',
            **oee_sum,
            'avg_efficiency': round(float(sp_agg['avg_eff'] or 0), 1),
            'total_breakage': sp_agg['total_breakage'] or 0,
            'total_output_kg': float(sp_agg['total_output'] or 0),
            'batch_count': sp_agg['batch_count'],
        })

    # ── بوبین‌پیچی OEE ──────────────────────────────────────────
    wd_machines = Machine.objects.filter(status='active', machine_type='winding')
    if line:
        wd_machines = wd_machines.filter(production_line=line)

    for m in wd_machines:
        wd_agg = WindingProd.objects.filter(
            machine=m, production_date__range=(date_from, date_to)
        ).aggregate(
            batch_count   = Count('id'),
            avg_eff       = Avg('efficiency_pct'),
            avg_cuts      = Avg('cuts_per_100km'),
            total_output  = Sum('output_weight_kg'),
            total_input   = Sum('input_weight_kg'),
            total_waste   = Sum('waste_weight_kg'),
        )
        avg_eff   = float(wd_agg['avg_eff']  or 0)
        avg_cuts  = float(wd_agg['avg_cuts'] or 0)
        # OEE کیفیت WD: cuts < 20 = 100%, cuts > 80 = 20%، خطی بین این دو
        quality_score = max(20.0, min(100.0, 100 - (avg_cuts - 20) * (80 / 60))) if avg_cuts > 20 else 100.0
        # OEE Performance از efficiency مستقیم
        performance  = avg_eff
        # Availability: داده توقف برای WD در حال حاضر از DowntimeLog
        downtime_min = DowntimeLog.objects.filter(
            machine=m, start_time__date__range=(date_from, date_to)
        ).aggregate(total=Sum('duration_min'))['total'] or 0
        days = max(1, (date_to - date_from).days + 1)
        total_planned_min = days * 8 * 60  # ۸ ساعت/روز
        availability = max(0.0, min(100.0,
            (total_planned_min - downtime_min) / total_planned_min * 100
        ))
        oee = round(availability * performance * quality_score / 10000, 1)
        machine_oee.append({
            'code': m.code, 'name': m.name, 'section': 'بوبین‌پیچی',
            'availability': round(availability, 1),
            'performance':  round(performance, 1),
            'quality':      round(quality_score, 1),
            'oee':          oee,
            'avg_efficiency': round(avg_eff, 1),
            'total_breakage': 0,
            'total_output_kg': float(wd_agg['total_output'] or 0),
            'batch_count': wd_agg['batch_count'],
            'extra': {'avg_cuts_per_100km': round(avg_cuts, 1)},
        })

    # ── دولاتابی OEE ────────────────────────────────────────────
    tfo_machines = Machine.objects.filter(status='active', machine_type='tfo')
    if line:
        tfo_machines = tfo_machines.filter(production_line=line)

    for m in tfo_machines:
        tfo_agg = TFOProd.objects.filter(
            machine=m, production_date__range=(date_from, date_to)
        ).aggregate(
            batch_count      = Count('id'),
            avg_eff          = Avg('efficiency_pct'),
            total_breakage   = Sum('breakage_count'),
            total_output     = Sum('output_weight_kg'),
        )
        avg_eff      = float(tfo_agg['avg_eff']       or 0)
        total_brk    = tfo_agg['total_breakage'] or 0
        batch_count  = tfo_agg['batch_count'] or 1
        # OEE کیفیت TFO: breakage < 5/بچ = 100%، > 20/بچ = 50%
        brk_per_batch = total_breakage / batch_count if (total_breakage := total_brk) else 0
        quality_score = max(50.0, min(100.0, 100 - (brk_per_batch - 5) * (50 / 15))) if brk_per_batch > 5 else 100.0
        performance   = avg_eff
        downtime_min  = DowntimeLog.objects.filter(
            machine=m, start_time__date__range=(date_from, date_to)
        ).aggregate(total=Sum('duration_min'))['total'] or 0
        days = max(1, (date_to - date_from).days + 1)
        total_planned_min = days * 8 * 60
        availability = max(0.0, min(100.0,
            (total_planned_min - downtime_min) / total_planned_min * 100
        ))
        oee = round(availability * performance * quality_score / 10000, 1)
        machine_oee.append({
            'code': m.code, 'name': m.name, 'section': 'دولاتابی',
            'availability': round(availability, 1),
            'performance':  round(performance, 1),
            'quality':      round(quality_score, 1),
            'oee':          oee,
            'avg_efficiency': round(avg_eff, 1),
            'total_breakage': total_brk,
            'total_output_kg': float(tfo_agg['total_output'] or 0),
            'batch_count': tfo_agg['batch_count'],
            'extra': {'breakage_per_batch': round(brk_per_batch, 1)},
        })

    # ── هیت‌ست OEE ──────────────────────────────────────────────
    hs_machines = Machine.objects.filter(status='active', machine_type='heatset')
    if line:
        hs_machines = hs_machines.filter(production_line=line)

    for m in hs_machines:
        hs_agg = HeatsetBatch.objects.filter(
            machine=m, production_date__range=(date_from, date_to)
        ).aggregate(
            batch_count  = Count('id'),
            pass_count   = Count('id', filter=Q(quality_result='pass')),
            fail_count   = Count('id', filter=Q(quality_result='fail')),
            total_kg     = Sum('batch_weight_kg'),
            avg_duration = Avg('duration_min'),
        )
        total_b = hs_agg['batch_count'] or 0
        pass_count = hs_agg['pass_count'] or 0
        # OEE کیفیت HS = نرخ قبولی مستقیم
        quality_score = round(pass_count / total_b * 100, 1) if total_b > 0 else 0.0
        # Performance HS: نسبت مدت واقعی / مدت استاندارد (فرض ۱۲۰ دقیقه استاندارد)
        avg_dur = float(hs_agg['avg_duration'] or 0)
        standard_cycle = 120  # دقیقه
        performance = min(100.0, round(standard_cycle / avg_dur * 100, 1)) if avg_dur > 0 else 0.0
        downtime_min = DowntimeLog.objects.filter(
            machine=m, start_time__date__range=(date_from, date_to)
        ).aggregate(total=Sum('duration_min'))['total'] or 0
        days = max(1, (date_to - date_from).days + 1)
        total_planned_min = days * 8 * 60
        availability = max(0.0, min(100.0,
            (total_planned_min - downtime_min) / total_planned_min * 100
        ))
        oee = round(availability * performance * quality_score / 10000, 1)
        machine_oee.append({
            'code': m.code, 'name': m.name, 'section': 'هیت‌ست',
            'availability': round(availability, 1),
            'performance':  round(performance, 1),
            'quality':      quality_score,
            'oee':          oee,
            'avg_efficiency': quality_score,   # pass rate نمایش
            'total_breakage': hs_agg['fail_count'] or 0,
            'total_output_kg': float(hs_agg['total_kg'] or 0),
            'batch_count': total_b,
            'extra': {'pass_rate': quality_score, 'avg_cycle_min': round(avg_dur, 1)},
        })

    machine_oee.sort(key=lambda x: x['oee'], reverse=True)

    # ── trend راندمان روزانه (همه بخش‌ها) ─────────────────────
    trend_data = []
    current = date_from
    while current <= date_to:
        sp_day  = _filter_by_line(SpinningProd.objects.filter(production_date=current, status='completed'), line)
        wd_day  = _filter_by_line(WindingProd.objects.filter(production_date=current), line)
        tfo_day = _filter_by_line(TFOProd.objects.filter(production_date=current), line)
        trend_data.append({
            'date':       _jstr(current),
            'ring':       float(sp_day.aggregate(a=Avg('efficiency_pct'))['a']  or 0),
            'winding':    float(wd_day.aggregate(a=Avg('efficiency_pct'))['a']  or 0),
            'tfo':        float(tfo_day.aggregate(a=Avg('efficiency_pct'))['a'] or 0),
        })
        current += timedelta(days=1)

    # OEE میانگین کل
    oee_avg = round(sum(m['oee'] for m in machine_oee) / len(machine_oee), 1) if machine_oee else 0

    # بهترین OEE هر بخش برای KPI card
    def _best(section):
        items = [m for m in machine_oee if m.get('section') == section and m['batch_count'] > 0]
        return max(items, key=lambda x: x['oee']) if items else None

    ctx.update({
        'machine_oee':   machine_oee,
        'oee_avg':       oee_avg,
        'best_ring':     _best('رینگ'),
        'best_winding':  _best('بوبین‌پیچی'),
        'best_tfo':      _best('دولاتابی'),
        'best_heatset':  _best('هیت‌ست'),
        'date_from':     date_from,
        'date_to':       date_to,
        'chart_data':    json.dumps([{'label': m['code'], 'value': m['oee'], 'section': m.get('section', '')} for m in machine_oee]),
        'trend_data':    json.dumps(trend_data),
        'page_title':    'گزارش راندمان و OEE',
    })
    return render(request, 'reports/oee_report.html', ctx)


# ═══════════════════════════════════════════════════════════════
# R2.1 مقایسه بازه زمانی
# ═══════════════════════════════════════════════════════════════

def _period_stats(date_from, date_to, line=None):
    stats = {}
    sp_qs = _filter_by_line(SpinningProd.objects.filter(production_date__range=(date_from, date_to), status='completed'), line)
    sp = sp_qs.aggregate(total_output=Sum('output_weight'), total_input=Sum('input_weight'),
                         avg_efficiency=Avg('efficiency_pct'), total_breakage=Sum('breakage_count'), batch_count=Count('id'))
    stats.update({'spinning_output': float(sp['total_output'] or 0), 'spinning_efficiency': round(float(sp['avg_efficiency'] or 0), 1),
                  'spinning_breakage': sp['total_breakage'] or 0, 'spinning_batches': sp['batch_count']})

    bl_qs = _filter_by_line(BlowroomBatch.objects.filter(production_date__range=(date_from, date_to), status='completed'), line)
    bl = bl_qs.aggregate(total_waste=Sum('waste_weight'), total_input=Sum('total_input_weight'))
    bl_in  = float(bl['total_input'] or 0); bl_wst = float(bl['total_waste'] or 0)
    stats['blowroom_waste_pct'] = round(bl_wst / bl_in * 100, 2) if bl_in > 0 else 0

    dt_qs = _filter_by_line(DowntimeLog.objects.filter(start_time__date__range=(date_from, date_to)), line)
    dt = dt_qs.aggregate(total_min=Sum('duration_min'), count=Count('id'), total_loss=Sum('production_loss'))
    stats.update({'downtime_min': dt['total_min'] or 0, 'downtime_count': dt['count'], 'production_loss': float(dt['total_loss'] or 0)})

    from apps.orders.models import Order
    stats['orders_delivered'] = Order.objects.filter(status='delivered', updated_at__date__range=(date_from, date_to)).count()
    return stats


@login_required
def compare_periods(request):
    ctx  = _report_base_context(request)
    line = ctx['selected_line']
    today = date.today()
    try:
        a_from = date.fromisoformat(request.GET.get('a_from', ''))
        a_to   = date.fromisoformat(request.GET.get('a_to', ''))
    except (ValueError, TypeError):
        a_from = today - timedelta(days=14); a_to = today - timedelta(days=8)
    try:
        b_from = date.fromisoformat(request.GET.get('b_from', ''))
        b_to   = date.fromisoformat(request.GET.get('b_to', ''))
    except (ValueError, TypeError):
        b_from = today - timedelta(days=7); b_to = today

    period_a = _period_stats(a_from, a_to, line)
    period_b = _period_stats(b_from, b_to, line)
    comparisons = []
    metrics = [('تولید رینگ (kg)', 'spinning_output', 'kg', True),
               ('راندمان رینگ (%)', 'spinning_efficiency', '%', True),
               ('پارگی نخ', 'spinning_breakage', 'بار', False),
               ('ضایعات حلاجی (%)', 'blowroom_waste_pct', '%', False),
               ('توقفات (دقیقه)', 'downtime_min', 'دقیقه', False),
               ('سفارشات تحویلی', 'orders_delivered', '', True)]
    for label, key, unit, higher_is_better in metrics:
        val_a = period_a.get(key, 0); val_b = period_b.get(key, 0)
        change_pct = round((val_b - val_a) / abs(val_a) * 100, 1) if val_a else 0
        comparisons.append({'label': label, 'unit': unit, 'val_a': val_a, 'val_b': val_b,
                            'change_pct': change_pct, 'is_better': (change_pct > 0) == higher_is_better,
                            'is_same': change_pct == 0})
    ctx.update({'a_from': a_from, 'a_to': a_to, 'b_from': b_from, 'b_to': b_to, 'comparisons': comparisons,
                'chart_data': json.dumps({'labels': [c['label'] for c in comparisons[:6]],
                                          'period_a': [c['val_a'] for c in comparisons[:6]],
                                          'period_b': [c['val_b'] for c in comparisons[:6]]}),
                'page_title': 'مقایسه بازه زمانی'})
    return render(request, 'reports/compare_periods.html', ctx)


# ═══════════════════════════════════════════════════════════════
# R2.2 عملکرد اپراتورها
# ═══════════════════════════════════════════════════════════════

@login_required
def operator_performance(request):
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']
    sp_qs = _filter_by_line(SpinningProd.objects.filter(production_date__range=(date_from, date_to), status='completed'), line)
    by_operator = sp_qs.values('operator__first_name', 'operator__last_name', 'operator__username').annotate(
        total_output=Sum('output_weight'), avg_efficiency=Avg('efficiency_pct'),
        total_breakage=Sum('breakage_count'), batch_count=Count('id')).order_by('-total_output')
    operators = [{'name': (f"{o['operator__first_name']} {o['operator__last_name']}".strip() or o['operator__username']),
                  'total_output': float(o['total_output'] or 0), 'avg_efficiency': round(float(o['avg_efficiency'] or 0), 1),
                  'total_breakage': o['total_breakage'] or 0, 'batch_count': o['batch_count']} for o in by_operator]
    by_shift = sp_qs.values('shift__name', 'shift__code').annotate(
        total_output=Sum('output_weight'), avg_efficiency=Avg('efficiency_pct'),
        total_breakage=Sum('breakage_count'), batch_count=Count('id')).order_by('-total_output')
    shifts = [{'name': s['shift__name'] or s['shift__code'], 'total_output': float(s['total_output'] or 0),
               'avg_efficiency': round(float(s['avg_efficiency'] or 0), 1),
               'total_breakage': s['total_breakage'] or 0, 'batch_count': s['batch_count']} for s in by_shift]
    ctx.update({'operators': operators, 'shifts': shifts, 'date_from': date_from, 'date_to': date_to,
                'chart_operators': json.dumps([{'label': o['name'][:15], 'value': o['total_output']} for o in operators[:10]]),
                'chart_shifts':    json.dumps([{'label': s['name'], 'value': s['avg_efficiency']} for s in shifts]),
                'page_title': 'عملکرد اپراتورها'})
    return render(request, 'reports/operator_performance.html', ctx)


# ═══════════════════════════════════════════════════════════════
# D.2 گزارش بوبین‌پیچی
# ═══════════════════════════════════════════════════════════════

@login_required
def winding_report(request):
    """گزارش تخصصی بوبین‌پیچی: cuts, کیفیت، راندمان."""
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']
    qs   = _filter_by_line(
        WindingProd.objects.filter(production_date__range=(date_from, date_to))
                           .select_related('machine', 'operator', 'shift'), line
    )
    agg = qs.aggregate(
        total_batches=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        total_input_kg=Sum('input_weight_kg'),
        total_output_kg=Sum('output_weight_kg'),
        total_waste_kg=Sum('waste_weight_kg'),
        avg_cuts=Avg('cuts_per_100km'),
        avg_splices=Avg('splices_per_100km'),
        avg_efficiency=Avg('efficiency_pct'),
    )
    total_in = float(agg['total_input_kg'] or 0)
    total_waste = float(agg['total_waste_kg'] or 0)
    waste_pct = round(total_waste / total_in * 100, 2) if total_in > 0 else 0

    grade_dist = {
        'A (< 20)':   qs.filter(cuts_per_100km__lt=20).count(),
        'B (20-40)':  qs.filter(cuts_per_100km__gte=20, cuts_per_100km__lt=40).count(),
        'C (40-60)':  qs.filter(cuts_per_100km__gte=40, cuts_per_100km__lt=60).count(),
        'D (≥ 60)':   qs.filter(cuts_per_100km__gte=60).count(),
    }
    by_machine = qs.filter(cuts_per_100km__isnull=False).values('machine__code').annotate(
        avg_cuts=Avg('cuts_per_100km'), avg_eff=Avg('efficiency_pct'),
        total_output=Sum('output_weight_kg'), count=Count('id')).order_by('avg_cuts')[:10]

    trend_data = []
    current = date_from
    while current <= date_to:
        d_qs = qs.filter(production_date=current)
        d_agg = d_qs.aggregate(avg_cuts=Avg('cuts_per_100km'), avg_eff=Avg('efficiency_pct'), total_kg=Sum('output_weight_kg'))
        trend_data.append({'date': _jstr(current), 'cuts': round(float(d_agg['avg_cuts'] or 0), 1),
                           'efficiency': round(float(d_agg['avg_eff'] or 0), 1),
                           'kg': round(float(d_agg['total_kg'] or 0), 1)})
        current += timedelta(days=1)

    ctx.update({
        'agg': agg, 'waste_pct': waste_pct, 'grade_dist': grade_dist,
        'by_machine': by_machine, 'recent_batches': qs.order_by('-production_date', '-created_at')[:15],
        'trend_data': json.dumps(trend_data),
        'grade_chart': json.dumps([{'label': k, 'value': v} for k, v in grade_dist.items()]),
        'date_from': date_from, 'date_to': date_to, 'page_title': 'گزارش بوبین‌پیچی',
    })
    return render(request, 'reports/winding_report.html', ctx)


# ═══════════════════════════════════════════════════════════════
# D.3 گزارش دولاتابی
# ═══════════════════════════════════════════════════════════════

@login_required
def tfo_report(request):
    """گزارش تخصصی دولاتابی: TPM، پارگی، راندمان."""
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']
    qs   = _filter_by_line(
        TFOProd.objects.filter(production_date__range=(date_from, date_to))
                       .select_related('machine', 'operator', 'shift'), line
    )
    agg = qs.aggregate(
        total_batches=Count('id'),
        completed=Count('id', filter=Q(status='completed')),
        total_input_kg=Sum('input_weight_kg'),
        total_output_kg=Sum('output_weight_kg'),
        total_waste_kg=Sum('waste_weight_kg'),
        avg_tpm=Avg('twist_tpm'),
        avg_efficiency=Avg('efficiency_pct'),
        total_breakage=Sum('breakage_count'),
    )
    total_in = float(agg['total_input_kg'] or 0)
    total_waste = float(agg['total_waste_kg'] or 0)
    waste_pct = round(total_waste / total_in * 100, 2) if total_in > 0 else 0

    twist_dist = {
        'S — چپ‌تاب': qs.filter(twist_direction='S').count(),
        'Z — راست‌تاب': qs.filter(twist_direction='Z').count(),
    }
    by_machine = qs.values('machine__code').annotate(
        avg_eff=Avg('efficiency_pct'), avg_breakage=Avg('breakage_count'),
        avg_tpm=Avg('twist_tpm'), total_output=Sum('output_weight_kg'), count=Count('id')
    ).order_by('-avg_eff')[:10]

    trend_data = []
    current = date_from
    while current <= date_to:
        d_agg = qs.filter(production_date=current).aggregate(
            avg_eff=Avg('efficiency_pct'), total_breakage=Sum('breakage_count'), total_kg=Sum('output_weight_kg'))
        trend_data.append({'date': _jstr(current), 'efficiency': round(float(d_agg['avg_eff'] or 0), 1),
                           'breakage': d_agg['total_breakage'] or 0, 'kg': round(float(d_agg['total_kg'] or 0), 1)})
        current += timedelta(days=1)

    ctx.update({
        'agg': agg, 'waste_pct': waste_pct, 'twist_dist': twist_dist,
        'by_machine': by_machine, 'recent_batches': qs.order_by('-production_date', '-created_at')[:15],
        'trend_data': json.dumps(trend_data),
        'twist_chart': json.dumps([{'label': k, 'value': v} for k, v in twist_dist.items()]),
        'date_from': date_from, 'date_to': date_to, 'page_title': 'گزارش دولاتابی TFO',
    })
    return render(request, 'reports/tfo_report.html', ctx)


# ═══════════════════════════════════════════════════════════════
# D.4 گزارش هیت‌ست
# ═══════════════════════════════════════════════════════════════

@login_required
def heatset_report(request):
    """گزارش تخصصی هیت‌ست: pass/fail، دما، آنکاژ."""
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']
    qs   = _filter_by_line(
        HeatsetBatch.objects.filter(production_date__range=(date_from, date_to))
                            .select_related('machine', 'operator', 'shift'), line
    )
    agg = qs.aggregate(
        total_batches=Count('id'),
        pass_count=Count('id', filter=Q(quality_result='pass')),
        fail_count=Count('id', filter=Q(quality_result='fail')),
        conditional_count=Count('id', filter=Q(quality_result='conditional')),
        total_weight_kg=Sum('batch_weight_kg'),
        avg_temp=Avg('temperature_c'),
        avg_duration=Avg('duration_min'),
        avg_shrinkage=Avg('shrinkage_pct'),
    )
    total = agg['total_batches'] or 1
    pass_rate = round((agg['pass_count'] or 0) / total * 100, 1)

    by_fiber = qs.values('fiber_type').annotate(
        count=Count('id'), pass_count=Count('id', filter=Q(quality_result='pass')),
        avg_temp=Avg('temperature_c'), avg_shrinkage=Avg('shrinkage_pct')).order_by('-count')

    stability_dist = {
        'عالی':  qs.filter(twist_stability='excellent').count(),
        'خوب':   qs.filter(twist_stability='good').count(),
        'متوسط': qs.filter(twist_stability='fair').count(),
        'ضعیف':  qs.filter(twist_stability='poor').count(),
    }

    trend_data = []
    current = date_from
    while current <= date_to:
        d_qs   = qs.filter(production_date=current)
        d_tot  = d_qs.count()
        d_pass = d_qs.filter(quality_result='pass').count()
        d_agg  = d_qs.aggregate(avg_temp=Avg('temperature_c'))
        trend_data.append({'date': _jstr(current), 'total': d_tot, 'pass': d_pass,
                           'fail': d_tot - d_pass,
                           'pass_rate': round(d_pass / d_tot * 100, 1) if d_tot > 0 else 0,
                           'avg_temp': round(float(d_agg['avg_temp'] or 0), 1)})
        current += timedelta(days=1)

    ctx.update({
        'agg': agg, 'pass_rate': pass_rate, 'by_fiber': by_fiber,
        'stability_dist': stability_dist,
        'recent_batches': qs.order_by('-production_date', '-created_at')[:15],
        'trend_data': json.dumps(trend_data),
        'quality_chart': json.dumps([
            {'label': 'قبول',    'value': agg['pass_count'] or 0},
            {'label': 'رد',      'value': agg['fail_count'] or 0},
            {'label': 'مشروط',   'value': agg['conditional_count'] or 0},
        ]),
        'stability_chart': json.dumps([{'label': k, 'value': v} for k, v in stability_dist.items()]),
        'date_from': date_from, 'date_to': date_to, 'page_title': 'گزارش هیت‌ست',
    })
    return render(request, 'reports/heatset_report.html', ctx)


# ═══════════════════════════════════════════════════════════════
# D.5 گزارش زنجیره تولید
# ═══════════════════════════════════════════════════════════════

@login_required
def production_chain(request):
    """گزارش زنجیره SP→WD→TFO→HS: خلاصه تولید + ضایعات تجمعی."""
    date_from, date_to = _parse_date_range(request)
    ctx  = _report_base_context(request)
    line = ctx['selected_line']

    def _lf(qs):
        return qs.filter(production_line=line) if line else qs

    # ── رینگ ──────────────────────────────────────────────────────
    sp_qs  = _lf(SpinningProd.objects.filter(production_date__range=(date_from, date_to)))
    sp_agg = sp_qs.aggregate(total=Count('id'), completed=Count('id', filter=Q(status='completed')),
                              total_kg=Sum('output_weight'), avg_eff=Avg('efficiency_pct'))

    # ── بوبین‌پیچی ─────────────────────────────────────────────────
    wd_qs  = _lf(WindingProd.objects.filter(production_date__range=(date_from, date_to)))
    wd_agg = wd_qs.aggregate(total=Count('id'), completed=Count('id', filter=Q(status='completed')),
                              total_input_kg=Sum('input_weight_kg'), total_kg=Sum('output_weight_kg'),
                              waste_kg=Sum('waste_weight_kg'), avg_eff=Avg('efficiency_pct'),
                              avg_cuts=Avg('cuts_per_100km'))

    # ── دولاتابی ───────────────────────────────────────────────────
    tfo_qs  = _lf(TFOProd.objects.filter(production_date__range=(date_from, date_to)))
    tfo_agg = tfo_qs.aggregate(total=Count('id'), completed=Count('id', filter=Q(status='completed')),
                                total_kg=Sum('output_weight_kg'), waste_kg=Sum('waste_weight_kg'),
                                avg_eff=Avg('efficiency_pct'), total_breakage=Sum('breakage_count'))

    # ── هیت‌ست ─────────────────────────────────────────────────────
    hs_qs  = _lf(HeatsetBatch.objects.filter(production_date__range=(date_from, date_to)))
    hs_agg = hs_qs.aggregate(total=Count('id'), pass_count=Count('id', filter=Q(quality_result='pass')),
                               fail_count=Count('id', filter=Q(quality_result='fail')),
                               total_kg=Sum('batch_weight_kg'), avg_temp=Avg('temperature_c'))
    hs_t = hs_agg['total'] or 1
    hs_agg['pass_rate'] = round((hs_agg['pass_count'] or 0) / hs_t * 100, 1)

    # ── ضایعات زنجیره ──────────────────────────────────────────────
    sp_kg  = float(sp_agg['total_kg']         or 0)
    wd_in  = float(wd_agg['total_input_kg']   or 0)
    wd_out = float(wd_agg['total_kg']         or 0)
    tfo_out = float(tfo_agg['total_kg']       or 0)
    hs_out  = float(hs_agg['total_kg']        or 0)

    def _loss_row(stage, inp, out):
        loss = max(0, round(inp - out, 1))
        return {'stage': stage, 'input_kg': round(inp, 1), 'output_kg': round(out, 1),
                'loss_kg': loss, 'loss_pct': round(loss / inp * 100, 1) if inp > 0 else 0}

    chain_losses = [
        _loss_row('رینگ → WD (ورودی)',   sp_kg,  wd_in),
        _loss_row('WD — ضایعات داخلی',   wd_in,  wd_out),
        _loss_row('WD → TFO',            wd_out, tfo_out),
        _loss_row('TFO → HS',            tfo_out, hs_out),
    ]

    # ── داده نمودار ستونی (خروجی هر مرحله) ────────────────────────
    stage_chart = json.dumps([
        {'label': 'رینگ',        'value': round(sp_kg, 1)},
        {'label': 'WD خروجی',   'value': round(wd_out, 1)},
        {'label': 'TFO خروجی',  'value': round(tfo_out, 1)},
        {'label': 'HS خروجی',   'value': round(hs_out, 1)},
    ])

    # ── روند روزانه همه مراحل ──────────────────────────────────────
    trend_data = []
    current = date_from
    while current <= date_to:
        sp_d  = _lf(SpinningProd.objects.filter(production_date=current)).aggregate(kg=Sum('output_weight'))['kg'] or 0
        wd_d  = _lf(WindingProd.objects.filter(production_date=current)).aggregate(kg=Sum('output_weight_kg'))['kg'] or 0
        tfo_d = _lf(TFOProd.objects.filter(production_date=current)).aggregate(kg=Sum('output_weight_kg'))['kg'] or 0
        hs_d  = _lf(HeatsetBatch.objects.filter(production_date=current)).aggregate(kg=Sum('batch_weight_kg'))['kg'] or 0
        trend_data.append({'date': _jstr(current),
                           'sp': round(float(sp_d), 1), 'wd': round(float(wd_d), 1),
                           'tfo': round(float(tfo_d), 1), 'hs': round(float(hs_d), 1)})
        current += timedelta(days=1)

    ctx.update({
        'sp_agg': sp_agg, 'wd_agg': wd_agg, 'tfo_agg': tfo_agg, 'hs_agg': hs_agg,
        'chain_losses': chain_losses,
        'stage_chart': stage_chart,
        'trend_data': json.dumps(trend_data),
        'date_from': date_from, 'date_to': date_to,
        'page_title': 'گزارش زنجیره تولید',
    })
    return render(request, 'reports/production_chain.html', ctx)


# ═══════════════════════════════════════════════════════════════
# Excel exports
# ═══════════════════════════════════════════════════════════════

def _excel_response(filename):
    r = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    r['Content-Disposition'] = f'attachment; filename={filename}'
    return r


def _excel_header(ws, headers, color):
    from openpyxl.styles import Font, Alignment, PatternFill
    hf = PatternFill(start_color=color, end_color=color, fill_type='solid')
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hf
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[chr(64 + col)].width = 16


@login_required
def export_production_excel(request):
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse('openpyxl نصب نیست', status=500)
    date_from, date_to = _parse_date_range(request)
    qs = SpinningProd.objects.filter(production_date__range=(date_from, date_to)).select_related('machine', 'operator', 'shift').order_by('production_date')
    wb = Workbook(); ws = wb.active; ws.title = 'تولید رینگ'; ws.sheet_view.rightToLeft = True
    headers = ['شماره بچ', 'تاریخ', 'ماشین', 'اپراتور', 'شیفت', 'نمره نخ', 'تاب', 'دوک', 'ورودی(kg)', 'خروجی(kg)', 'پارگی', 'راندمان%', 'وضعیت']
    _excel_header(ws, headers, '4361EE')
    for row, p in enumerate(qs, 2):
        jd = jdatetime.date.fromgregorian(date=p.production_date)
        vals = [p.batch_number, jd.strftime('%Y/%m/%d'), p.machine.code, p.operator.get_full_name(),
                str(p.shift or ''), p.yarn_count, float(p.twist_tpm or 0), p.num_spindles_active,
                float(p.input_weight or 0), float(p.output_weight or 0), p.breakage_count,
                float(p.efficiency_pct or 0), p.get_status_display()]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    response = _excel_response(f'spinning_{date_from}_{date_to}.xlsx')
    wb.save(response)
    return response


@login_required
def export_winding_excel(request):
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse('openpyxl نصب نیست', status=500)
    date_from, date_to = _parse_date_range(request)
    qs = WindingProd.objects.filter(production_date__range=(date_from, date_to)).select_related('machine', 'operator', 'shift').order_by('production_date')
    wb = Workbook(); ws = wb.active; ws.title = 'بوبین‌پیچی'; ws.sheet_view.rightToLeft = True
    headers = ['شماره بچ', 'تاریخ', 'ماشین', 'اپراتور', 'شیفت', 'ورودی(kg)', 'خروجی(kg)', 'ضایعات(kg)', 'برش/100km', 'اتصال/100km', 'راندمان%', 'نوع بوبین', 'وضعیت']
    _excel_header(ws, headers, '0891B2')
    for row, p in enumerate(qs, 2):
        jd = jdatetime.date.fromgregorian(date=p.production_date)
        vals = [p.batch_number, jd.strftime('%Y/%m/%d'), p.machine.code, p.operator.get_full_name(),
                str(p.shift or ''), float(p.input_weight_kg or 0), float(p.output_weight_kg or 0),
                float(p.waste_weight_kg or 0), p.cuts_per_100km, p.splices_per_100km,
                float(p.efficiency_pct or 0), p.get_package_type_display(), p.get_status_display()]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    response = _excel_response(f'winding_{date_from}_{date_to}.xlsx')
    wb.save(response)
    return response


@login_required
def export_heatset_excel(request):
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse('openpyxl نصب نیست', status=500)
    date_from, date_to = _parse_date_range(request)
    qs = HeatsetBatch.objects.filter(production_date__range=(date_from, date_to)).select_related('machine', 'operator', 'shift').order_by('production_date')
    wb = Workbook(); ws = wb.active; ws.title = 'هیت‌ست'; ws.sheet_view.rightToLeft = True
    headers = ['شماره بچ', 'تاریخ', 'ماشین', 'اپراتور', 'نوع دستگاه', 'نوع الیاف', 'دما(°C)', 'فشار(bar)', 'مدت(min)', 'وزن(kg)', 'تعداد بوبین', 'نتیجه', 'آنکاژ%', 'پایداری تاب']
    _excel_header(ws, headers, 'DC2626')
    for row, b in enumerate(qs, 2):
        jd = jdatetime.date.fromgregorian(date=b.production_date)
        vals = [b.batch_number, jd.strftime('%Y/%m/%d'), b.machine.code, b.operator.get_full_name(),
                b.get_machine_type_hs_display(), b.get_fiber_type_display(),
                float(b.temperature_c), float(b.steam_pressure_bar or 0),
                b.duration_min, float(b.batch_weight_kg), b.packages_count,
                b.get_quality_result_display() if b.quality_result else '',
                float(b.shrinkage_pct or 0), b.get_twist_stability_display() if b.twist_stability else '']
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    response = _excel_response(f'heatset_{date_from}_{date_to}.xlsx')
    wb.save(response)
    return response


@login_required
def export_downtime_excel(request):
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse('openpyxl نصب نیست', status=500)
    date_from, date_to = _parse_date_range(request)
    qs = DowntimeLog.objects.filter(start_time__date__range=(date_from, date_to)).select_related('machine', 'operator', 'shift').order_by('start_time')
    wb = Workbook(); ws = wb.active; ws.title = 'توقفات'; ws.sheet_view.rightToLeft = True
    headers = ['ماشین', 'تاریخ شروع', 'مدت(min)', 'دسته دلیل', 'جزئیات', 'تلفات(kg)', 'اپراتور', 'شیفت']
    _excel_header(ws, headers, 'DC3545')
    for row, d in enumerate(qs, 2):
        jdt_str = jdatetime.datetime.fromgregorian(datetime=d.start_time).strftime('%Y/%m/%d %H:%M') if d.start_time else ''
        vals = [d.machine.code, jdt_str, d.duration_min, d.get_reason_category_display(),
                d.reason_detail, float(d.production_loss or 0), d.operator.get_full_name(), str(d.shift or '')]
        for col, v in enumerate(vals, 1):
            ws.cell(row=row, column=col, value=v)
    response = _excel_response(f'downtime_{date_from}_{date_to}.xlsx')
    wb.save(response)
    return response
